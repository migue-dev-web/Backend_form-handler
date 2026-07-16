import io
import pandas as pd
from sqlalchemy.orm import Session
from fastapi.responses import StreamingResponse
from fastapi import HTTPException
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import app.models as models

def generar_excel_consolidado(formularios_ids: list[int], db: Session):
    # 1. Buscar los registros de los formularios en la Base de Datos
    formularios = db.query(models.FormularioDB).filter(models.FormularioDB.id.in_(formularios_ids)).all()
    
    if not formularios:
        raise HTTPException(status_code=404, detail="No se encontraron formularios válidos.")

    # 2. Crear un archivo Excel en memoria (BytesIO) para no ocupar almacenamiento en Render
    output = io.BytesIO()
    
    # Inicializar el escritor de pandas con el motor openpyxl
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        
        for form in formularios:
            if not form.sheet_id:
                continue # Saltar si el formulario no tiene link/ID asociado
            
            # Construir la URL de exportación directa a CSV
            url_csv = f"https://docs.google.com/spreadsheets/d/{form.sheet_id}/export?format=csv"
            
            try:
                # Pandas descarga y parsea el CSV directamente desde internet en una sola línea
                df = pd.read_csv(url_csv)
                
                # Limpiar el nombre del formulario para usarlo como pestaña (máximo 31 caracteres, regla de Excel)
                nombre_pestaña = form.nombre[:30].replace("/", "-").replace("\\", "-")
                
                # Escribir el DataFrame en su respectiva pestaña
                # Empezamos en la fila 4 para dejar espacio a un encabezado estético
                df.to_excel(writer, sheet_name=nombre_pestaña, index=False, startrow=3)
                
                # ---- ESTILIZADO PROFESIONAL CON OPENPYXL ----
                workbook = writer.book
                worksheet = writer.sheets[nombre_pestaña]
                worksheet.views.sheetView[0].showGridLines = True # Asegurar que las cuadrículas se vean
                
                # Estilos de diseño
                azul_marino = "1F497D"
                gris_claro = "F2F5F8"
                
                fuente_titulo = Font(name="Arial", size=14, bold=True, color=azul_marino)
                fuente_headers = Font(name="Arial", size=10, bold=True, color="FFFFFF")
                fuente_datos = Font(name="Arial", size=10)
                
                fill_header = PatternFill(start_color=azul_marino, end_color=azul_marino, fill_type="solid")
                fill_zebra = PatternFill(start_color=gris_claro, end_color=gris_claro, fill_type="solid")
                
                border_fino = Border(
                    left=Side(style='thin', color='D9D9D9'),
                    right=Side(style='thin', color='D9D9D9'),
                    top=Side(style='thin', color='D9D9D9'),
                    bottom=Side(style='thin', color='D9D9D9')
                )
                
                # Añadir Título de la sección en la fila 1
                worksheet["A1"] = f"REPORTE CONSOLIDADO: {form.nombre.upper()}"
                worksheet["A1"].font = fuente_titulo
                worksheet["A2"] = f"ID de origen: {form.sheet_id}"
                worksheet["A2"].font = Font(name="Arial", size=9, italic=True, color="595959")
                
                # Estilizar los encabezados de la tabla (Fila 4 de Excel)
                num_columnas = df.shape[1]
                for col in range(1, num_columnas + 1):
                    cell = worksheet.cell(row=4, column=col)
                    cell.font = fuente_headers
                    cell.fill = fill_header
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                worksheet.row_dimensions[4].height = 25
                
                # Estilizar las filas de datos (Zebra striping y bordes)
                num_filas = df.shape[0]
                for row in range(5, num_filas + 5):
                    worksheet.row_dimensions[row].height = 18
                    for col in range(1, num_columnas + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.font = fuente_datos
                        cell.border = border_fino
                        if row % 2 == 0:
                            cell.fill = fill_zebra
                
                # Autoajustar el ancho de las columnas según el contenido para evitar que se corte el texto
                for col in worksheet.columns:
                    max_len = max(len(str(cell.value or '')) for cell in col)
                    col_letter = get_column_letter(col[0].column)
                    worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
                    
            except Exception as e:
                print(f" Error al procesar el Sheet {form.sheet_id}: {e}")
                # Podrías crear una pestaña de error si un documento no es accesible
                continue

    # 3. Preparar el puntero del archivo en memoria para ser transmitido por HTTP
    output.seek(0)
    
    # 4. Retornar una respuesta de tipo Stream para que el navegador inicie la descarga inmediata
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=reporte_formularios_consolidado.xlsx"}
    )

def generar_pdf_consolidado(formularios_ids: list[int], db: Session):
    formularios = db.query(models.FormularioDB).filter(models.FormularioDB.id.in_(formularios_ids)).all()
    if not formularios:
        raise HTTPException(status_code=404, detail="No se encontraron formularios válidos.")
    
    output = io.BytesIO()
    doc = SimpleDocTemplate(
        output, pagesize=letter,
        rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40
    )
    
    styles = getSampleStyleSheet()
    style_titulo = ParagraphStyle('T1', parent=styles['Heading1'], fontName='Helvetica-Bold', fontSize=20, textColor=colors.HexColor("#1F497D"), spaceAfter=15, alignment=TA_CENTER)
    style_sub = ParagraphStyle('S1', parent=styles['Normal'], fontName='Helvetica-Oblique', fontSize=10, textColor=colors.HexColor("#595959"), spaceAfter=25, alignment=TA_CENTER)
    style_sec = ParagraphStyle('Sec1', parent=styles['Heading2'], fontName='Helvetica-Bold', fontSize=14, textColor=colors.HexColor("#1F497D"), spaceBefore=15, spaceAfter=10)
    style_h = ParagraphStyle('H1', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=9, textColor=colors.white, alignment=TA_CENTER)
    style_d = ParagraphStyle('D1', parent=styles['Normal'], fontName='Helvetica', fontSize=9, textColor=colors.black, alignment=TA_LEFT)

    story = []
    story.append(Paragraph("REPORTE CONSOLIDADO DE FORMULARIOS", style_titulo))
    story.append(Paragraph("Documento PDF generado a partir de datos integrados en Google Sheets", style_sub))
    story.append(Spacer(1, 10))

    for form in formularios:
        if not form.sheet_id:
            continue
            
        url_csv = f"https://docs.google.com/spreadsheets/d/{form.sheet_id}/export?format=csv"
        
        try:
            df = pd.read_csv(url_csv)
            if df.empty:
                continue
                
            story.append(Paragraph(f"Formulario: {form.nombre}", style_sec))
            
            data_tabla = []
            headers = [Paragraph(str(col), style_h) for col in df.columns]
            data_tabla.append(headers)
            
            for index, row in df.iterrows():
                fila_parrafos = [Paragraph(str(val) if pd.notna(val) else "", style_d) for val in row]
                data_tabla.append(fila_parrafos)
            
            col_widths = [530 / len(df.columns)] * len(df.columns)
            tabla_pdf = Table(data_tabla, colWidths=col_widths, repeatRows=1)
            
            estilo_tabla = TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor("#1F497D")),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('LEFTPADDING', (0, 0), (-1, -1), 6),
                ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                ('INNERGRID', (0, 0), (-1, -1), 0.5, colors.HexColor("#D9D9D9")),
                ('BOX', (0, 0), (-1, -1), 0.5, colors.HexColor("#1F497D")),
            ])
            
            for i in range(1, len(data_tabla)):
                if i % 2 == 0:
                    estilo_tabla.add('BACKGROUND', (0, i), (-1, i), colors.HexColor("#F2F5F8"))
            
            tabla_pdf.setStyle(estilo_tabla)
            story.append(tabla_pdf)
            story.append(Spacer(1, 20))
            
        except Exception as e:
            print(f" Error en PDF para el Sheet {form.sheet_id}: {e}")
            continue

    doc.build(story)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=reporte_formularios.pdf"}
    )